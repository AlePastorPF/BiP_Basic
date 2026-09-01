#!/usr/bin/env python3
"""
build_dashboard.py
-------------------
Regenera dist/index.html (el dashboard interactivo) a partir de:
  - data/SIC_BiP_Full.xlsx   (planilla "Ball in Play")
  - assets/logos/*           (escudos)
  - crest_map.json           (que partido usa que escudo / nombre)
  - dashboard/template.html  (la interfaz, no se toca en cada corrida)

Uso:
    python3 build_dashboard.py
    python3 build_dashboard.py --excel data/otro_archivo.xlsx --out dist/index.html

No requiere nada raro: openpyxl y Pillow (ambos se instalan con
`pip install -r requirements.txt`).
"""
import argparse
import base64
import io
import json
import statistics
import sys
from pathlib import Path

import openpyxl
from PIL import Image

ROOT = Path(__file__).parent
DEFAULT_EXCEL = ROOT / "data" / "SIC_BiP_Full.xlsx"
DEFAULT_TEMPLATE = ROOT / "dashboard" / "template.html"
DEFAULT_OUT = ROOT / "dist" / "index.html"
DEFAULT_LOGOS_DIR = ROOT / "assets" / "logos"
DEFAULT_CREST_MAP = ROOT / "crest_map.json"

SHEET_NAME = "Ball in Play_2026"  # cambiar si la solapa cambia de nombre/temporada


def load_crest_map(path: Path) -> dict:
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    raw.pop("_comment", None)
    return raw


def humanize(code: str) -> str:
    """Fallback si un partido no está en crest_map.json: SIC_Ejemplo_2 -> 'Ejemplo (V2)'."""
    name = code.replace("SIC_", "").replace("_", " ")
    if name.endswith(" 2"):
        name = name[:-2].strip() + " (V2)"
    return name


def encode_logos(logos_dir: Path, needed_keys: set) -> dict:
    out = {}
    exts = [".png", ".jpg", ".jpeg", ".webp"]
    for key in needed_keys:
        if key is None:
            continue
        found = None
        for ext in exts:
            p = logos_dir / f"{key}{ext}"
            if p.exists():
                found = p
                break
        if not found:
            print(f"  [!] No se encontró escudo para '{key}' en {logos_dir} — se omite.", file=sys.stderr)
            continue
        img = Image.open(found).convert("RGBA")
        img.thumbnail((120, 120), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        out[key] = base64.b64encode(buf.getvalue()).decode("ascii")
    return out


def aggregate(excel_path: Path, crest_map: dict):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(
            f"No encontré la solapa '{SHEET_NAME}' en {excel_path}. "
            f"Solapas disponibles: {wb.sheetnames}. "
            f"Editá SHEET_NAME en build_dashboard.py si cambió el nombre."
        )
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    partido_order, seen = [], set()
    for r in rows:
        p = r[0]
        if p is not None and p not in seen:
            seen.add(p)
            partido_order.append(p)

    player_pos = {}
    players = set()
    for r in rows:
        partido, _, jugador, seq, pos, dur, dist, acel, hsr, big, cont = r
        if jugador is None:
            continue
        players.add(jugador)
        player_pos[jugador] = pos
    players = sorted(players)

    agg = {}
    for r in rows:
        partido, _, jugador, seq, pos, dur, dist, acel, hsr, big, cont = r
        if partido is None or jugador is None:
            continue
        key = (partido, jugador)
        a = agg.setdefault(key, {"big_vals": [], "dist": 0.0, "dur": 0, "acel": 0, "hsr": 0.0, "cont": 0, "events": 0})
        a["events"] += 1
        a["dur"] += dur or 0
        a["dist"] += dist or 0
        a["acel"] += acel or 0
        a["hsr"] += hsr or 0
        a["cont"] += cont or 0
        if big is not None:
            a["big_vals"].append(big)

    matches = {}
    for partido in partido_order:
        entry = crest_map.get(partido, {})
        rival = entry.get("rival", humanize(partido))
        logo = entry.get("logo")

        match_players, all_big = [], []
        for j in players:
            key = (partido, j)
            if key not in agg:
                continue
            a = agg[key]
            avg_big = round(statistics.mean(a["big_vals"]), 2) if a["big_vals"] else None
            if avg_big is not None:
                all_big.append(avg_big)
            match_players.append({
                "jugador": j.strip(),
                "posicion": player_pos[j],
                "avgBiG": avg_big,
                "bigCount": len(a["big_vals"]),
                "distTotal": round(a["dist"], 1),
                "durTotal": a["dur"],
                "acelAlta": a["acel"],
                "hsrTotal": round(a["hsr"], 2),
                "contactos": a["cont"],
                "events": a["events"],
            })
        team_avg = round(statistics.mean(all_big), 2) if all_big else None
        matches[partido] = {
            "rival": rival,
            "logo": logo,
            "teamAvgBiG": team_avg,
            "players": match_players,
        }

    data = {
        "partidoOrder": partido_order,
        "players": [p.strip() for p in players],
        "matches": matches,
    }

    needed_logo_keys = {m["logo"] for m in matches.values()} | {"SIC"}
    return data, needed_logo_keys


def build(excel_path: Path, template_path: Path, out_path: Path, logos_dir: Path, crest_map_path: Path):
    print(f"→ Leyendo {excel_path}")
    crest_map = load_crest_map(crest_map_path)
    data, needed_logo_keys = aggregate(excel_path, crest_map)
    print(f"  {len(data['partidoOrder'])} partidos, {len(data['players'])} jugadores")

    unmapped = [p for p in data["partidoOrder"] if p not in crest_map]
    if unmapped:
        print(f"  [!] Partidos sin escudo asignado en crest_map.json (uso nombre genérico): {unmapped}",
              file=sys.stderr)

    print(f"→ Codificando escudos desde {logos_dir}")
    logos = encode_logos(logos_dir, needed_logo_keys)

    data_js = json.dumps(data, ensure_ascii=False)
    logos_js = json.dumps(logos)
    payload = f"window.__SIC_DATA__ = {data_js};\nwindow.__SIC_LOGOS__ = {logos_js};\n"

    print(f"→ Generando {out_path}")
    html = template_path.read_text(encoding="utf-8")
    marker = '<script src="payload.js"></script>'
    if marker not in html:
        raise SystemExit(f"No encontré el marcador {marker!r} en {template_path}. ¿Se editó la plantilla?")
    html = html.replace(marker, f"<script>\n{payload}\n</script>")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")
    print(f"✔ Listo: {out_path} ({out_path.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Regenera el dashboard SIC Ball in Play")
    ap.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument("--logos", type=Path, default=DEFAULT_LOGOS_DIR)
    ap.add_argument("--crest-map", type=Path, default=DEFAULT_CREST_MAP)
    args = ap.parse_args()

    build(args.excel, args.template, args.out, args.logos, args.crest_map)
